import os
from openpyxl import Workbook

# Create a new Excel file
wb = Workbook()
ws = wb.active
ws.title = "Repository Sizes"

# Set the column headers
ws['A1'] = "Repository"
ws['B1'] = "Size"

# Initialize the row counter
row = 2

# Iterate through the directory
for filename in os.listdir("path/to/directory"):
    if filename.endswith(".txt"):
        # Open the text file and read the values
        with open(os.path.join("path/to/directory", filename), 'r') as f:
            lines = f.readlines()
            repository = lines[0].strip()
            size = lines[1].strip()

        # Write the values to the Excel file
        ws[f'A{row}'] = repository
        ws[f'B{row}'] = size
        row += 1

# Save the Excel file
wb.save("repository_sizes.xlsx")